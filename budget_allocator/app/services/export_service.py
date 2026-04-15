"""
app/services/export_service.py
------------------------------
Responsible for generating an in-memory Bytes buffer natively utilizing `openpyxl`.
"""

import io
from typing import List, Dict, Any
import openpyxl
from pydantic import BaseModel

def generate_excel_export(data: List[Any], sheet_name: str = "Export") -> io.BytesIO:
    """
    Dynamically generate an in-memory Excel Workbook explicitly bounding
    the Keys natively across schemas mapped in the payload.
    """
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if not data:
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    # Setup styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    currency_format = '"$"#,##0.00'

    # Hardcoded vertical labels mapped exactly to the screenshot
    labels = [
        ("TC Count", "tc_count", False),
        ("Start Date", "start_date", False),
        ("End Date", "end_date", False),
        ("Manual TC Count", "manual_tc_count", True),
        ("Automation TC Count", "automation_tc_count", False),
        ("Adhoc Request", "adhoc_request", True),
        ("Total TC", "total_tc", False),
        ("Duration in Days", "duration_in_days", False),
        ("Duration wks", "duration_wks", False),
        ("Manual HC", "manual_hc", False),
        ("Automation HC", "automation_hc", False),
        ("Manual HC cost", "manual_hc_cost", True),
        ("Automation HC cost", "automation_hc_cost", False),
        ("Lead Cost", "lead_cost", True),
        ("SQPM Cost of Boise 70%", "sqpm_cost_boise", False),
        ("PL-50%", "pl_cost", True),
        ("Per WQE - 40%", "per_wqe_cost", False),
        ("aSQPM - 80%", "asqpm_cost", True),
        ("Lab Techician & Manager - 40%", "lab_tech_manager_cost", False),
        ("Project Manager - 40%", "project_manager_cost", True),
        ("", None, False), # Blank spacer row 22
        ("Total Budget", "total_budget", False),
    ]

    # Column A is the Team name. If there are multiple budgets, we'll just put the first team name,
    # or arguably we can just write the Team names dynamically.
    # To match the screenshot, Column A is A2:A23 merged "CPE".
    # Col B is labels.
    
    # Write Labels (Column B)
    sheet.cell(row=1, column=2, value="Walmart SKU Stage").font = header_font
    sheet.cell(row=1, column=2).fill = header_fill
    sheet.cell(row=1, column=2).alignment = center_align

    for idx, (label, key, is_gray) in enumerate(labels):
        row_idx = idx + 2
        cell = sheet.cell(row=row_idx, column=2, value=label)
        if is_gray:
            cell.fill = gray_fill
        if label == "Total Budget":
            cell.font = Font(bold=True)

    # Write Data Blocks (Columns C, D... E, F...)
    current_col = 3
    for item in data:
        # Header Row 1
        run_name = item.get("run_name", "Unknown Run")
        team_name = item.get("team_name", "Team")
        
        cpe_cell = sheet.cell(row=1, column=current_col, value=run_name)
        cpe_cell.font = header_font
        cpe_cell.fill = header_fill
        cpe_cell.alignment = center_align

        cmt_cell = sheet.cell(row=1, column=current_col + 1, value="Comments")
        cmt_cell.font = header_font
        cmt_cell.fill = header_fill
        cmt_cell.alignment = center_align

        # Write Values
        for idx, (label, key, is_gray) in enumerate(labels):
            row_idx = idx + 2
            val_cell = sheet.cell(row=row_idx, column=current_col)
            
            if is_gray:
                val_cell.fill = gray_fill
                
            if key is not None:
                val = item.get(key)
                
                # Format Dates
                if key in ["start_date", "end_date"] and val:
                    val_cell.value = str(val)
                # Format Currency
                elif key.endswith("_cost") or key == "total_budget":
                    val_cell.value = float(val) if val is not None else 0.0
                    val_cell.number_format = currency_format
                    if key == "total_budget":
                        val_cell.font = Font(bold=True)
                # Numeric formatting
                else:
                    if val is not None:
                        try:
                            val_cell.value = float(val)
                        except ValueError:
                            val_cell.value = str(val)

        # Merge Team Name into Column A for the rows of this block
        # If we have multiple, we just merge Column A vertically. 
        # (Assuming all budgets belong to the same team. If not, Col A might be weird, but we'll write it anyway).
        sheet.cell(row=2, column=1, value=team_name).alignment = center_align
        sheet.cell(row=2, column=1).font = Font(bold=True)
        # Merge A2:A23
        try:
            sheet.merge_cells(start_row=2, start_column=1, end_row=23, end_column=1)
        except Exception:
            pass # Already merged by previous block if multiple budgets

        current_col += 2

    # Auto-adjust column widths
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 30
    for col in range(3, current_col):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    workbook.save(buffer)
    buffer.seek(0)
    return buffer
